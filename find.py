"""find patterns in domains_info domains"""

import re

from pathlib import PurePath
from openpyxl import load_workbook, Workbook

from domains import read_domains_file


def write_domains_to_spreadsheet(filename: str, domains_patterns: dict) -> True:
    """Write domain information to spreadsheet

    :param domains_info: dictionary with domain names

    :return: domains_info
    """
    # Open workbook in memory
    workbook = Workbook()
    # Create a sheet named 'Domains'
    sheet = workbook.create_sheet(title='Domains')
    # For each pattern name write to a column
    for column, pattern_name in enumerate(domains_patterns):
        column = column + 1    # excel starts counting from 1
        # Write the name of the pattern in row 1
        sheet.cell(row=1, column=column).value = pattern_name
        # For each domain discovered in the search
        for d_count, domain in enumerate(sorted(domains_patterns[pattern_name]['domains'])):
            # Write the domain in subsequent rows (same column as above)
            sheet.cell(row=d_count+2, column=column).value = domain
    # Delete default Sheet
    del workbook["Sheet"]
    print("\nWriting to file")
    # Write workbook to file
    workbook.save(filename)


def match_pattern(domains_info):
    domains_patterns= {
        'cambridge ALL': {
            'patterns': [r"cam(b.*ge){0,1}(\w|\.|-)"],
            'domains': []},
        'ballsports': {
             'patterns': [r"ball.cam(b.*ge){0,1}(\w|\.|-)", r"(foot|basket|net){0,1}"],
             'domains': []},
        'athletics': {
             'patterns': [r"ath{0,1}letic(s){0,1}.cam(b.*ge){0,1}(\w|\.|-)"],
             'domains': []},
        'football': {
            'patterns': [r"(football|soccer).cam(b.*ge){0,1}(\w|\.|-)"],
            'domains':[]},
        }
    # Use the following as another list to track domains not included above
    domain_list = [d for d in domains_info]
    for patterns in domains_patterns:
        print(f"\tsearching pattern: {patterns}...")
        for domain in domains_info:
            for pattern in domains_patterns[patterns]['patterns']:
                if len(re.findall(pattern.lower(), domain.lower())) == 0:
                    pattern_found = False
                    break
                else:
                    pattern_found = True
                    # Remove domain found from the left over list
                    try:
                        domain_list.remove(domain)
                    except ValueError:
                        pass
            if pattern_found == True:
                domains_patterns[patterns]['domains'].append(domain)
    # Append the left over list as another column in the data sheet
    domains_patterns['ALL OTHER DOMAINS'] = {'domains': domain_list}
    return domains_patterns

def domains_info_find():
    """Match with regular expression

    :return: True
    """
    in_filename = PurePath('data','domains_list.csv')
    out_filename = PurePath('output', 'domain_patterns.xlsx')
    print(f"Opening '{in_filename}'")
    domains_info = read_domains_file(in_filename)
    print(f"Searching patterns:")
    domains_patterns = match_pattern(domains_info)
    print(f"Writing results to '{out_filename}'")
    write_domains_to_spreadsheet(out_filename, domains_patterns)
    print("Done!!")


#---------------------------------------------------------------


if __name__ == "__main__":
    domains_info_find()
