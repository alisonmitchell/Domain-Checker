"""Domain explorer"""

from pathlib import PurePath
import sys
import platform
import socket
import datetime
import random
import csv

import click
from openpyxl import load_workbook, Workbook
# for Linux use:  pip install pywhois,
# for Windows use: pip install python-whois
import whois


def get_whois(domains_info: dict, n_retries: int) -> dict:
    """Access whois data for a given domain
    https://pypi.org/project/whois/

    :param domains_info: dictionary with domain names
    :param n_retries: number of retries for whois function

    :return: domains_info
    """

    def _iter_over_domains(Domains: list, domains_info: dict) -> tuple:
        # Note that some servers refuse to answer too many queries, by
        # shuffling the domains_info, we ensure probability of calling the same
        # server is reduced. Also whois.query uses slow_down of 2sec
        random.shuffle(Domains)
        domains_timeout = []
        for domain in Domains:
            # add dictionary properties for domain
            domains_info[domain]['registrar'] = ""
            domains_info[domain]['expiration_date'] = ""
            domains_info[domain]['last_update'] = ""
            domains_info[domain]['name'] = ""

            try:
                # it should work for both Linux and Windows
                if platform.system() == 'Linux':
                    whois_data = whois.query(domain, slow_down=2)
                elif platform.system() == 'Windows':
                    whois_data = whois.whois(domain)
                else:
                    raise OSError("OS not supported")
            except:
                # Domain name not found by whois, try next
                domains_timeout.append(domain)
                domains_info[domain]['WHOIS'] = "timeout"
                continue
            domains_info[domain]['WHOIS'] = "Found"
            try:
                domains_info[domain]['registrar'] = f"{whois_data.registrar}"
            except AttributeError:
                pass
            try:
                domains_info[domain]['name'] = f"{whois_data.name}"
            except AttributeError:
                pass
            try:
                domains_info[domain]['expiration_date'] = f"{whois_data.expiration_date}"
            except AttributeError:
                pass
            try:
                domains_info[domain]['last_update'] = f"{whois_data.last_update}"
            except AttributeError:
                pass
        return (domains_timeout, domains_info)

    # iterate over domains in list and add to dictionary,
    # retry if whois query fails due to timeouts
    Domains = [d for d in domains_info]
    # Repeats until you have found a percentage of sites
    for _ in range(n_retries):
        Domains, domains_info = _iter_over_domains(Domains, domains_info)
    return domains_info


def check_host(domains_info: dict) -> dict:
    """Iterate through every entry in domains_info and find the IP address

    :param domains_info: dictionary with domain names

    :return: domains_info
    """
    # iterate over domains
    for domain in domains_info:
        # Only query for domains where registrar is known
        if domains_info[domain]['registrar'] == "":
            domains_info[domain]['IP address'] = 'N/A'
            continue
        try:
            domains_info[domain]['IP address'] = socket.gethostbyname(domain)
        except socket.gaierror:
            domains_info[domain]['IP address'] = 'NOT FOUND'
    return domains_info


def read_domains_file(domains_filename: str) -> dict:
    """Read all domains_info domains parameters from main csv file

    :param domains_filename: filename of csv

    :return: dictionary
    """
    domains_info = {}
    # Open the csv file
    with open(domains_filename, newline='') as csvfile:
        domains = csv.reader(csvfile)
        # step over every row
        for row_n, domain in enumerate(domains):
            if row_n == 0:
                continue
            domains_info[domain[0]] = {}
    return domains_info


def write_domains_to_spreadsheet(filename: str, domains_info: dict) -> True:
    """Write domain information to spreadsheet

    :param domains_info: dictionary with domain names

    :return: domains_info
    """
    workbook = Workbook()
    sheet = workbook.create_sheet(title='Domains')
    sheet.cell(row=1, column=1).value = 'Domain'
    sheet.cell(row=1, column=2).value = 'IP address'
    sheet.cell(row=1, column=3).value = 'Registrar'
    sheet.cell(row=1, column=4).value = 'Expiration date'
    sheet.cell(row=1, column=5).value = 'WHOIS'
    for count, domain in enumerate(domains_info):
        domain
        sheet.cell(row=count+2, column=1).value = domain
        sheet.cell(row=count+2, column=2).value = domains_info[domain]['IP address']
        sheet.cell(row=count+2, column=3).value = domains_info[domain]['registrar']
        sheet.cell(row=count+2, column=4).value =  \
                f"{domains_info[domain]['expiration_date']}"
        sheet.cell(row=count+2, column=5).value = domains_info[domain]['WHOIS']
    del workbook["Sheet"]
    print("\nWriting to file")
    workbook.save(filename)


def main(n_retries: int) -> True:
    """Main function to the program handling reading, processing and writing of the data

    :param n_retries: number of retries for the whois function

    :return: A dict with the spreadsheet content
    """
    in_filename = PurePath('data','domains_list.csv')
    out_filename = PurePath('output', 'domain_whois.xlsx')
    # Read the CSV file with domains.    
    domains_info = read_domains_file(in_filename)
    # Whois query DNS for domain's attributes (e.g. registrar) 
    domains_info = get_whois(domains_info, n_retries)
    # DNS-lookup for obtaining the IP address of the domain
    domains_info = check_host(domains_info)
    # Write domain info to file
    write_domains_to_spreadsheet(out_filename, domains_info)
    print('*** Done!***')
    return True


def get_arguments_with_click() -> True:
    """
    Handle arguments from Command Line using 'click'

    :return: True
    """
    ABOUT = """ Read in a csv file, iterate over a list of domains to access whois data, 
    add the domains to a dictionary, and output information to a spreadsheet"""
    # Creates an instance click called 'cli'
    @click.group(help=ABOUT)
    def cli():
        "Main call to the CLI"
        pass
    
    # 'Extend' the functionality of 'cli' with a new function called 'run_solver'
    @cli.command('run', help="Run the domain explorer function")
    @click.option('--n_retries', '-n', default=10,
                  help='Number of retries for whois call')
    def run_solver(n_retries):
        main(n_retries)
            
    cli()

    return True


#---------------------------------------------------------------


if __name__ == "__main__":
    get_arguments_with_click()
    